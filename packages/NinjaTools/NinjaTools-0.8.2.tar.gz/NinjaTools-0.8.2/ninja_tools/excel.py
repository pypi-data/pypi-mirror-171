import re
import warnings

try:
    import openpyxl
    import xlrd
    import webcolors
except ImportError:
    raise 'pip install ninjatools[excel] or ninjatools[all] to use excel functions!'

# Remove warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')


class ColorObject(object):
    def __init__(self, hex_, rgb, color_name):
        self.hex = hex_
        self.rgb = rgb
        self.name = color_name

    def __repr__(self):
        return self.rgb


class Excel:
    def __init__(self, workbook_path):
        self.wb_xlrd = xlrd.open_workbook(workbook_path)
        self.wb_openpyxl = openpyxl.load_workbook(workbook_path, data_only=True)
        self.sheet_name = None

    @staticmethod
    def get_cell(cell) -> tuple:
        """
        Returns a tuple of the cell
        :param cell:
        :return:
        """
        r = re.compile("([a-zA-Z]+)([0-9]+)")
        m = r.match(cell)
        col = m.group(1).upper()
        number_row = int(m.group(2)) - 1

        abc = ["A", "B", "C", "D", "E", "F", "G", "H", "I",
               "J", "K", "L", "M", "N", "O", "P", "Q", "R",
               "S", "T", "U", "V", "W", "X", "Y", "Z"]

        number_col = 0

        for char in col:
            number_col += abc.index(char)

        return number_row, number_col

    def get_sheets(self) -> list:
        """
        Returns a list of sheets
        :return:
        """
        return [_ for _ in self.wb_xlrd.sheet_names()]

    def cell(self, cell, sheet_name=None) -> str:
        """
        Reads a cell
        :param cell:
        :param sheet_name:
        :return:
        """
        sheet_name = sheet_name if sheet_name else self.sheet_name
        ws = self.wb_xlrd.sheet_by_name(sheet_name)
        return str(ws.cell(*self.get_cell(cell)).value)

    def get_color(self, cell) -> ColorObject:
        """
        Returns the color of a cell
        :param cell:
        :return:
        """
        ws = self.wb_openpyxl[self.sheet_name]
        hex_color = ws[cell].fill.start_color.index
        hex_color = f'#{hex_color[2:]}'
        rgb_color = webcolors.hex_to_rgb(hex_color)
        rgb_color = [rgb_color.red, rgb_color.green, rgb_color.blue]
        color_name = self.closest_color(rgb_color)

        return ColorObject(hex_color, rgb_color, color_name)

    def read_range(self, cell_1, cell_2, sheet_name=None) -> list:
        """
        Reads a range of cells
        :param cell_1:
        :param cell_2:
        :param sheet_name:
        :return:
        """
        sheet_name = sheet_name if sheet_name else self.sheet_name

        var = self.get_cell(cell_1)
        var2 = self.get_cell(cell_2)

        ws = self.wb_xlrd.sheet_by_name(sheet_name)

        data = []
        for _ in range(var[0], var2[0] + 1):
            temp = []
            for __ in range(var[1], var2[1] + 1):
                temp.append(ws.cell(_, __).value)

            data.append(temp)

        return data

    def read_indefinitely(self, start_cell: str, num_of_columns: int, num_of_rows=None, steps: int = 1) -> list:
        """
        Reads a range of cells indefinitely
        :param start_cell:
        :param num_of_columns:
        :param num_of_rows:
        :param steps:
        :return:
        """

        data = []

        ord_num = ord(start_cell[0])
        i = int(start_cell[1:])

        while True:
            try:
                data.append([self.cell(f'{chr(ord_num + _)}{i}') for _ in range(num_of_columns)])
                if num_of_rows:
                    if len(data) >= num_of_rows:
                        break
            except (Exception,):
                break
            i += steps

        return data

    @staticmethod
    def convert_to_dict(data: list, key_idx: int, header_idx: int = 0) -> dict:
        """
        Converts a list to a dictionary
        :param data:
        :param key_idx:
        :param header_idx:
        :return:
        """
        dict_data = {}
        for idx, row in enumerate(data):
            if idx > 0:
                dict_data[row[key_idx]] = dict(zip(data[header_idx], row))

        return dict_data

    @staticmethod
    def closest_color(rgb):
        differences = {}
        for color_hex, color_name in webcolors.CSS3_HEX_TO_NAMES.items():
            r, g, b = webcolors.hex_to_rgb(color_hex)
            differences[sum([(r - rgb[0]) ** 2,
                             (g - rgb[1]) ** 2,
                             (b - rgb[2]) ** 2])] = color_name
        return differences[min(differences.keys())]

    # TODO: Cell write values/formulas

import argparse
import os
import xml.etree.ElementTree as xm

import pandas as pd
import requests as rq
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter

from geo_to_hca.utils.entrez_client import call_efetch
from geo_to_hca.utils.handle_errors import NotFoundENA

STATUS_ERROR_CODE = 400

class SraUtils:

    @staticmethod
    def get_content(accessions,accession_type):
        if len(accessions) < 100:
            xml = SraUtils.request_info(accessions,accession_type=accession_type)
            size = 'small'
            return xml,size
        else:
            size = 'large'
            parts_list = SraUtils.split_list(accessions, n=100)
            xmls = []
            for p in range(0,len(parts_list)):
                xml = SraUtils.request_info(parts_list[p],accession_type=accession_type)
                xmls.append(xml)
            return xmls,size

    @staticmethod
    def pubmed_id(project_pubmed_id):
        pubmed_response = call_efetch(db='pubmed', accessions=project_pubmed_id, rettype='xml')
        return xm.fromstring(pubmed_response.content)

def fetch_bioproject(bioproject_accession: str):
    xml_content = SraUtils.srp_bioproject(bioproject_accession)
    bioproject_metadata = xml_content.find('DocumentSummary')
    try:
        project_name = bioproject_metadata.find("Project").find('ProjectDescr').find('Name').text
    except:
        print("no project name")
        project_name = None
    try:
        project_title = bioproject_metadata.find("Project").find('ProjectDescr').find('Title').text
    except:
        print("no project title")
        project_title = None
    try:
        project_description = bioproject_metadata.find("Project").find('ProjectDescr').find('Description').text
    except:
        project_description = ''
        print("no project description")
    project_publication = bioproject_metadata.find("Project").find('ProjectDescr').find('Publication')
    try:
        if project_publication.find('DbType').text == 'Pubmed' or project_publication.find('DbType').text == 'ePubmed':
            try:
                project_pubmed_id = project_publication.attrib['id']
            except:
                project_pubmed_id = project_publication.find('Reference').text
    except:
        print("No publication for project %s was found: searching project title in EuropePMC" % (bioproject_accession))
    if not project_publication or not project_pubmed_id:
        if project_title:
            print("project title is: %s" % (project_title))
            url = rq.get(f'https://www.ebi.ac.uk/europepmc/webservices/rest/search?query={project_title}')
            if url.status_code == STATUS_ERROR_CODE:
                raise NotFoundENA(url, project_title)
            else:
                xml_content = xm.fromstring(url.content)
                try:
                    results = list()
                    result_list = xml_content.find("resultList")
                    for result in result_list:
                        results.append(result)
                    journal_title = results[0].find("journalTitle").text
                    if not journal_title or journal_title == '':
                        project_pubmed_id = ''
                        print("no publication results for project title in ENA")
                    else:
                        answer = input("A publication title has been found: %s.\nIs this the publication title associated with the GEO accession? [y/n]: " % (journal_title))
                        if answer.lower() in ['y',"yes"]:
                            project_pubmed_id = results[0].find("pmid").text
                        else:
                            journal_title = results[1].find("journalTitle").text
                            if not journal_title or journal_title == '':
                                project_pubmed_id = ''
                                print("no publication results for project title in ENA")
                            else:
                                answer = input("An alternative publication title has been found: %s.\nIs this the publication title associated with the GEO accession? [y/n]: " % (journal_title))
                                if answer.lower() in ['y', "yes"]:
                                    project_pubmed_id = results[1].find("pmid").text
                                else:
                                    journal_title = results[2].find("journalTitle").text
                                    if not journal_title or journal_title == '':
                                        project_pubmed_id = ''
                                        print("no publication results for project title in ENA")
                                    else:
                                        answer = input("An alternative publication title has been found: %s.\nIs this the publication title associated with the GEO accession? [y/n]: " % (journal_title))
                                        if answer.lower() in ['y', "yes"]:
                                            project_pubmed_id = results[2].find("pmid").text
                                        else:
                                            project_pubmed_id = ''
                                            print("no publication results for project title in ENA")
                except:
                    print("no publication results for project title in ENA")
                    project_pubmed_id = ''
        if not project_pubmed_id or project_pubmed_id == '':
            if project_name:
                print("project name is %s:" % (project_name))
                url = rq.get(f'https://www.ebi.ac.uk/europepmc/webservices/rest/search?query={project_name}')
                if url.status_code == STATUS_ERROR_CODE:
                    raise NotFoundENA(url, project_name)
                else:
                    xml_content = xm.fromstring(url.content)
                    try:
                        results = list()
                        result_list = xml_content.find("resultList")
                        for result in result_list:
                            results.append(result)
                        journal_title = results[0].find("journalTitle").text
                        if not journal_title or journal_title == '':
                            project_pubmed_id = ''
                            print("no publication results for project name in ENA")
                        else:
                            answer = input("A publication title has been found: %s.\nIs this the publication title associated with the GEO accession? [y/n]: " % (journal_title))
                            if answer.lower() in ['y',"yes"]:
                                project_pubmed_id = results[0].find("pmid").text
                            else:
                                journal_title = results[1].find("journalTitle").text
                                if not journal_title or journal_title == '':
                                    project_pubmed_id = ''
                                    print("no publication results for project name in ENA")
                                else:
                                    answer = input("An alternative publication title has been found: %s.\nIs this the publication title associated with the GEO accession? [y/n]: " % (journal_title))
                                    if answer.lower() in ['y', "yes"]:
                                        project_pubmed_id = results[1].find("pmid").text
                                    else:
                                        journal_title = results[2].find("journalTitle").text
                                        if not journal_title or journal_title == '':
                                            project_pubmed_id = ''
                                            print("no publication results for project name in ENA")
                                        else:
                                            answer = input("An alternative publication title has been found: %s.\nIs this the publication title associated with the GEO accession? [y/n]: " % (journal_title))
                                            if answer.lower() in ['y', "yes"]:
                                                project_pubmed_id = results[2].find("pmid").text
                                            else:
                                                project_pubmed_id = ''
                                                print("no publication results for project name in ENA")
                    except:
                        print("no publication results for project name in ENA")
                        project_pubmed_id = ''
        if not project_pubmed_id or project_pubmed_id == '':
            project_title = ''
            project_name = ''
            project_pubmed_id = ''
    return project_name,project_title,project_description,project_pubmed_id

def fetch_pubmed(project_pubmed_id: str,iteration: int):
    xml_content = SraUtils.pubmed_id(project_pubmed_id)
    author_list = list()
    grant_list=  list()
    try:
        title = xml_content.find("PubmedArticle").find("MedlineCitation").find("Article").find("ArticleTitle").text
    except:
        title = ''
        if iteration == 1:
            print("no publication title found")
    try:
        authors = xml_content.find("PubmedArticle").find("MedlineCitation").find("Article").find("AuthorList")
    except:
        if iteration == 1:
            print("no authors found in SRA")
        try:
            url = rq.get(f'https://www.ebi.ac.uk/europepmc/webservices/rest/search?query={title}')
            if url.status_code == STATUS_ERROR_CODE:
                raise NotFoundENA(url, title)
            else:
                xml_content_2 = xm.fromstring(url.content)
            try:
                results = list()
                result_list = xml_content_2.find("resultList")
                for result in result_list:
                    results.append(result)
                author_string = results[0].find("authorString").text
            except:
                authors = None
                if iteration == 1:
                    print("no authors found in ENA")
        except:
            authors = None
            if iteration == 1:
                print("no authors found in ENA")
    if authors is not None:
        for author in authors:
            try:
                lastname = author.find("LastName").text
            except:
                lastname = ''
            try:
                forename = author.find("ForeName").text
            except:
                forename = ''
            try:
                initials = author.find("Initials").text
            except:
                initials = ''
            try:
                affiliation = author.find('AffiliationInfo').find("Affiliation").text
            except:
                affiliation = ''
            author_list.append([lastname,forename,initials,affiliation])
    try:
        grants = xml_content.find("PubmedArticle").find("MedlineCitation").find("Article").find("GrantList")
    except:
        grants = None
        if iteration == 1:
            print("no grants found in SRA or ENA")
    if grants:
        for grant in grants:
            try:
                id = grant.find("GrantID").text
            except:
                id = ''
            try:
                agency = grant.find("Agency").text
            except:
                agency = ''
            grant_list.append([id,agency])
    try:
        articles = xml_content.find('PubmedArticle').find('PubmedData').find('ArticleIdList')
        for article_id in articles:
            if "/" in article_id.text:
                article_doi_id = article_id.text
    except:
        article_doi_id = ''
        if iteration == 1:
            print("no publication doi found")
    return title,author_list,grant_list,article_doi_id

def parse_xml(xml_content):
    for experiment_package in xml_content.findall('EXPERIMENT_PACKAGE'):
        yield experiment_package

def get_empty_df(workbook,tab_name):
    sheet = workbook[tab_name]
    values = sheet.values
    empty_df = pd.DataFrame(values)
    cols = empty_df.loc[0,]
    tab = pd.DataFrame(columns=cols)
    return tab

def write_to_wb(workbook: Workbook, tab_name: str, tab_content: pd.DataFrame) -> None:
    """
    Write the tab to the active workbook.
    :param workbook: str
                     Workbook extracted from the template
    :param tab_name: str
                     Name of the tab that is being modified
    :param tab_content: str
                        Content of the tab

    :returns None
    """
    worksheet = workbook[tab_name]
    # If more than 1 series is being filled for the same spreadsheet, find the first unfilled row
    row_not_filled = 6
    while True:
        if worksheet[f'A{row_not_filled}'].value or worksheet[f'B{row_not_filled}'].value:
            row_not_filled += 1
        else:
            break

    for index, key in enumerate(worksheet[4]):
        if not key.value:
            break
        if key.value not in tab_content.keys():
            continue
        for i in range(len(tab_content[key.value])):
            worksheet[f"{get_column_letter(index + 1)}{i + row_not_filled}"] = list(tab_content[key.value])[i]


def get_project_publication_tab_xls(workbook,tab_name,project_pubmed_id):
    tab = get_empty_df(workbook,tab_name)
    title,author_list,grant_list,article_doi_id = fetch_pubmed(project_pubmed_id,iteration=1)
    name_list = list()
    for author in author_list:
        name = author[0] + ' ' + author[2] + "||"
        name_list.append(name)
    name_list = ''.join(name_list)
    name_list = name_list[:len(name_list)-2]
    tab = tab.append({'project.publications.authors':name_list,
                      'project.publications.title':title,
                      'project.publications.doi':article_doi_id,
                      'project.publications.pmid':project_pubmed_id,
                      'project.publications.url':''}, ignore_index=True)
    write_to_wb(workbook, tab_name, tab)


def get_project_contributors_tab_xls(workbook,tab_name,project_pubmed_id):
    tab = get_empty_df(workbook,tab_name)
    title, author_list, grant_list, article_doi_id = fetch_pubmed(project_pubmed_id,iteration=2)
    for author in author_list:
        name = author[1] + ',,' + list(author)[0]
        affiliation = author[3]
        tab = tab.append({'project.contributors.name':name,'project.contributors.institution':affiliation}, ignore_index=True)
    write_to_wb(workbook, tab_name, tab)


def get_project_funders_tab_xls(workbook,tab_name,project_pubmed_id):
    tab = get_empty_df(workbook,tab_name)
    title, author_list, grant_list, article_doi_id = fetch_pubmed(project_pubmed_id,iteration=3)
    for grant in grant_list:
        tab = tab.append({'project.funders.grant_id':grant[0],'project.funders.organization':grant[1]}, ignore_index=True)
    write_to_wb(workbook, tab_name, tab)


def empty_worksheet(worksheet):
    if worksheet['A6'].value or worksheet['B6'].value:
        return False
    return True

def delete_unused_worksheets(workbook: Workbook) -> None:
    """
    Delete unused sheets from the metadata spreadsheet and the linked protocols.

    :param workbook: Workbook
                     Workbook object containing the metadata spreadsheet being modified
    :returns None
    """

    for worksheet_name in OPTIONAL_TABS:
        current_worksheet = workbook[worksheet_name]
        if current_worksheet and empty_worksheet(current_worksheet):
            del workbook[worksheet_name]
            if worksheet_name in LINKINGS:
                for linked_sheet in LINKINGS[worksheet_name]:
                    if empty_worksheet(workbook[linked_sheet]):
                        del workbook[linked_sheet]

def list_str(values):
    if "," not in values:
        raise argparse.ArgumentTypeError("Argument list not valid: comma separated list required")
    return values.split(',')

def check_file(path):
    if not os.path.exists(path):
        raise argparse.ArgumentTypeError("file %s does not exist" % (path))
    try:
        df = pd.read_csv(path, sep="\t")
    except:
        raise argparse.ArgumentTypeError("file %s is not a valid format" % (path))
    try:
        geo_accession_list = list(df["accession"])
    except:
        raise argparse.ArgumentTypeError("accession list column not found in file %s" % (path))
    return geo_accession_list

def main():

    parser = argparse.ArgumentParser()
    parser.add_argument('--pubmed_id',type=str,help='pubmed id')
    parser.add_argument('--input_file',type=check_file,help='optional path to tab-delimited input .txt file')
    parser.add_argument('--nthreads',type=int,default=1,
                        help='number of multiprocessing processes to use')
    parser.add_argument('--template',default="docs/hca_template.xlsx",
                        help='path to an HCA spreadsheet template (xlsx)')
    parser.add_argument('--header_row',type=int,default=4,
                        help='header row with HCA programmatic names')
    parser.add_argument('--input_row1',type=int,default=6,
                        help='HCA metadata input start row')
    parser.add_argument('--output_dir',default='spreadsheets/',
                        help='path to output directory; if it does not exist, the directory will be created')
    parser.add_argument('--output_log',type=bool,default=True,
                        help='True/False: should the output result log be created')

    args = parser.parse_args()


    if not os.path.exists(args.template):
        print("path to HCA template file not found; will revert to default: docs/hca_template.xlsx")
        template = "docs/hca_template.xlsx"
    try:
        workbook = load_workbook(filename=args.template)
        template = args.template
    except:
        print("specified HCA template file is not valid xlsx; will revert to default: docs/hca_template.xlsx")
        template = "docs/hca_template.xlsx"

    if not os.path.exists(args.output_dir):
        os.mkdir(args.output_dir)

    # Get Project metadata
 
    try:
        # get Project - Publications metadata: fetch as many fields as is possible using the above metadata accessions
        get_project_publication_tab_xls(workbook,tab_name="Project - Publications",project_pubmed_id=args.pubmed_id)
    except AttributeError:
        print(f'Publication attribute error.')

    try:
        # get Project - Contributors metadata: fetch as many fields as is possible using the above metadata accessions
        get_project_contributors_tab_xls(workbook,tab_name="Project - Contributors",project_pubmed_id=args.pubmed_id)
    except AttributeError:
        print(f'Contributors attribute error.')

    try:
        # get Project - Funders metadata: fetch as many fields as is possible using the above metadata accessions
        get_project_funders_tab_xls(workbook,tab_name="Project - Funders",project_pubmed_id=args.pubmed_id)
    except AttributeError:
        print(f'Funders attribute error.')

    out_file = str(args.pubmed_id) + ".xlsx"

    # Done
    print(f"Done. Saving workbook to excel file")
    workbook.save(out_file)

if __name__ == "__main__":
    main()

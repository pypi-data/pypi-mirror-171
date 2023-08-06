"""Модуль считывателя из формата xlsx."""

import datetime
from typing import Optional, Iterable

import openpyxl

from .base_reader import BaseReader


class ExcelReader(BaseReader):
    """Считыватель из формата xlsx."""

    def __init__(self, path: str, sheet_name: Optional[str] = None):
        """Конструктор считывателя из формата xlsx.

        :param path: путь к файлу
        :param sheet_name: название рабочего листа
        """
        super().__init__(path)
        self.work_book = openpyxl.load_workbook(path)
        self.sheet = self.work_book.active if sheet_name is None else self.work_book[sheet_name]
        self._headers: list[str] = [self.sheet.cell(1, j + 1).value for j in range(self.sheet.max_column)]

    @property
    def items(self) -> Iterable:
        """Перечисляем элементы с двух. Первая строка заголовок."""
        for i in range(2, self.sheet.max_row + 1):
            yield self.tree_transform({
                self._headers[j]: self.get_value(self.sheet.cell(i, j + 1).value)
                for j in range(self.sheet.max_column)
                if self._headers[j] is not None
            })

    @staticmethod
    def get_value(value):
        """Получение правильного значения ячейки."""
        if type(value) == str:
            return value.strip()
        elif type(value) == datetime.datetime or type(value) == datetime.date:
            return value.strftime('%Y-%m-%d')
        return value

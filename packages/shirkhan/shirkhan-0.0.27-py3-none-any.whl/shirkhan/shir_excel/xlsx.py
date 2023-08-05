from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
import os


class xlsx:
    def __init__(self, file, *args, **kwargs):
        pass
        # type:Workbook
        self.book = load_workbook(file, *args, **kwargs)

    @property
    def active(self):
        return self.book.active

    @property
    def active_sheet_name(self):
        return self.sheet_names()[0]

    def sheet_names(self):
        """
        返回sheet names列表
        :return:
        """
        return self.book.sheetnames

    def sheets(self):
        """
        返回sheets对象列表
        :return:
        """
        return self.book.worksheets

    def sheet_by_name(self, sheet_name: str = None) -> Worksheet:
        """
        通过名字获取sheet对象
        :param sheet_name:
        :return :
        """
        sheet_name = self.active_sheet_name if sheet_name is None else sheet_name
        return self.book[sheet_name]

    def sheet_by_index(self, sheet_index) -> Worksheet:
        """
        通过下标获取sheet对象
        :param sheet_index:
        :return:
        """
        return self.book.sheetnames[sheet_index]

    def rows_data(self, sheet_name=None):
        """
        返回转换数据后的data,而不是execl 对象
        :param sheet_name:
        :return:
        """
        rows = []
        for row in self.rows(sheet_name):
            item = []
            for cell in row:
                item.append(cell.value)
            rows.append(item)
        return rows

    def rows(self, sheet_name=None):
        """
        返回指定sheet名下的所有行
        :param sheet_name:
        :return:
        """
        sheet_name = self.active_sheet_name if sheet_name is None else sheet_name
        ws = self.book[sheet_name]  # type:Worksheet
        rows = ws.rows  # type:list[tuple[Cell]]
        return rows

    def rows_counts(self, sheet_name: str = None):
        """
        返回指定sheet有多少行
        :param sheet_name:
        :return:
        """
        sheet_name = self.active_sheet_name if sheet_name is None else sheet_name
        ws = self.book[sheet_name]  # type:Worksheet
        return ws.max_row

    def row_values(self, sheet_name: str = None, rowx: int = 1):
        """
        返回指定区域的行列表
        :param sheet_name:
        :param rowx:
        :return:
        """
        sheet_name = self.active_sheet_name if sheet_name is None else sheet_name
        ws = self.book[sheet_name]  # type:Worksheet
        cells = ws[rowx]  # type:tuple[Cell]
        return cells

    def row_types(self, sheet_name: str = None, rowx: int = 1):
        """
        返回指定行的 字段类型的数值列表
        :param sheet_name:
        :param rowx:
        :return:
        """
        sheet_name = self.active_sheet_name if sheet_name is None else sheet_name
        return [item.data_type for item in self.row_values(sheet_name, rowx)]

    def cols_counts(self, sheet_name: str = None):
        """
        返回指定sheet有多少列
        :param sheet_name:
        :return:
        """
        sheet_name = self.active_sheet_name if sheet_name is None else sheet_name
        return self.sheet_by_name(sheet_name).max_column

    def col_values(self, sheet_name: str = None, colx=1):
        """
            返回指定区域的列列表
        :param sheet_name:
        :param colx:
        :return list[Cell]:
        """
        sheet_name = self.active_sheet_name if sheet_name is None else sheet_name
        for item in self.rows(sheet_name):
            yield item[colx]

    def col_types(self, sheet_name: str = None, colx: int = 1, ):
        """
        返回指定列的 字段类型数值列表
        :param sheet_name:
        :param colx:
        :return:
        """
        sheet_name = self.active_sheet_name if sheet_name is None else sheet_name
        for item in self.col_values(sheet_name, colx):
            yield item.data_type

    def cell(self, sheet_name: str = None, rowx: int = 1, colx: int = 1) -> Cell:
        """
       通过行和列下标获取 cell 对象
        :param sheet_name:
        :param rowx:
        :param colx:
        :return:
        """
        sheet_name = self.active_sheet_name if sheet_name is None else sheet_name
        return self.sheet_by_name(sheet_name).cell(row=rowx, column=colx)

    def cell_value(self, sheet_name: str = None, rowx: int = 1, colx: int = 1):
        """
        返回指定sheet名下的指定下标的cell的值
        :param sheet_name:
        :param rowx:
        :param colx:
        :return:
        """
        sheet_name = self.active_sheet_name if sheet_name is None else sheet_name
        return self.cell(sheet_name, rowx, colx).value

    @staticmethod
    def load_workbook(file, *args, **kwargs):
        """
        打开excel 返回实例
        :param file:
        :param args:
        :param kwargs:
        :return:
        """
        return xlsx(file, *args, **kwargs)

    @staticmethod
    def write(file_path: str, sheet_rows: list):
        """
        rows创建excel，并保存给定目录下
        如果文件已经存在会抛出 assert，不然会创建整个目录和文件
        :param file_path:
        :param {str, list} sheet_rows_dict: {"sheetname":"rows"}
        :return:
        """

        assert os.path.exists(file_path) is False, f"文件 {file_path} 已存在"

        wb = Workbook()
        ws = wb.create_sheet(wb.sheetnames[0], 0)
        for row in sheet_rows:
            ws.append(row)

        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        wb.save(file_path)

        return True

    @staticmethod
    def write_by_dict(file_path: str, sheet_rows_dict: dict):
        """
        按照给定字典的key为sheet,value为rows创建excel，并保存给定目录下
        如果文件已经存在会抛出 assert，不然会创建整个目录和文件
        :param file_path:
        :param {str, list} sheet_rows_dict: {"sheetname":"rows"}
        :return:
        """

        assert os.path.exists(file_path) is False, f"文件 {file_path} 已存在"

        wb = Workbook()
        for index, item in enumerate(sheet_rows_dict.items()):
            sheet, rows = item
            ws = wb.create_sheet(sheet, index)
            for row in rows:
                ws.append(row)

        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        wb.save(file_path)

        return True


if __name__ == '__main__':
    pass
    # book = xlsx.load_workbook("aaa.xlsx")
    # print(book.row_values())
    # # print(book[ss])
    # print(book.sheets())
    # print(book.sheet_by_index(1))
    # print(book.rows(sheet_name))
    # print(book.rows_counts(sheet_name))
    # print(book.row_values(sheet_name, 100))
    # print(book.row_types(sheet_name, 1))
    # print(book.cols_counts(sheet_name))
    # print("aaa", *book.col_values(sheet_name, 1))
    # print("aaa", *book.col_types(sheet_name, 1))
    # print("ccc", book.cell(sheet_name, 1, 1))
    # print("ccca", book.cell_value(sheet_name, 1, 2))
    # xlsx.write("test/aaa.xlsx", [(1, 2)])
    # xlsx.write_by_dict("test/aaa.xlsx", {"shir": [(1, 2)]})
